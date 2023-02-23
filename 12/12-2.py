from openpyxl import load_workbook

xl_file_name = 'namelist.xlsx'

wb = load_workbook(xl_file_name)
ws = wb.active

name_list = []
birth_list = []
num_list = []

for i in range(1, ws.max_row+1):
    name_list.append(ws.cell(i, 1).value)
    birth_list.append(ws.cell(i,2).value)
    num_list.append(ws.cell(i,3).value)

print(name_list, birth_list, num_list)