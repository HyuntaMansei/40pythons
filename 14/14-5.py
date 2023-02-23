from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

smtp_name = 'smtp.naver.com'
smtp_port = 587

sender_email = "@naver.com"
sender_pwd = ''

msg = MIMEMultipart()
msg['Subject'] = '대량으로 보내는 메일'
msg['From'] = sender_email

text = """
대량으로 보내는 메일입니다.
폐기하시면 됩니다.
"""
msg.attach(MIMEText(text))

wbook = load_workbook('email_list.xlsx', data_only=True)
wsheet = wbook.active

s = smtplib.SMTP(smtp_name, smtp_port)
s.starttls()
s.login(sender_email, sender_pwd)

for i in range(1, wsheet.max_row+1):
    recver_email = wsheet.cell(i, 1).value
    print(recver_email)
    msg['To'] = recver_email

    s.sendmail(sender_email, recver_email, msg.as_string())

s.quit()