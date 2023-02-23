import requests
import re
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

headers = {
    'User-Agent': "Mozilla/5.0",
    'Content-Type': 'text/html; charset=utf-8'
}
url = 'https://m.blog.naver.com/act41/221578376673'

os.chdir(os.path.dirname(os.path.abspath(__file__)))

response = requests.get(url, headers=headers)
results = re.findall(r'[\w.-]+@[\w.-]+', response.text)
print(results)

try:
    wb = load_workbook('email_list.xlsx', data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
i = 1
j = 1

for r in results:
    sheet.cell(row=i, column=j).value = r
    i += 1
    j += 1
# for r in results:
#     sheet.append([r])

# sheet.append(results)

wb.save('email_list.xlsx')